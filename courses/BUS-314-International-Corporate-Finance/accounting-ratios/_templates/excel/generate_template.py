import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# Color fills
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

bold = Font(bold=True)
header_font = Font(bold=True, color='FFFFFF', size=11)
section_font = Font(bold=True, size=12)

# ========== BALANCE SHEET TAB ==========
ws_bal = wb.active
ws_bal.title = 'Balance Sheet'
ws_bal.column_dimensions['A'].width = 4
ws_bal.column_dimensions['B'].width = 42
ws_bal.column_dimensions['C'].width = 18
ws_bal.column_dimensions['D'].width = 18
ws_bal.column_dimensions['E'].width = 5
ws_bal.column_dimensions['F'].width = 42
ws_bal.column_dimensions['G'].width = 18
ws_bal.column_dimensions['H'].width = 18

ws_bal['B1'] = 'ASSETS'
ws_bal['B1'].font = section_font
ws_bal['F1'] = 'LIABILITIES & EQUITY'
ws_bal['F1'].font = section_font

ws_bal['C2'] = 'Current Year'
ws_bal['D2'] = 'Prior Year'
ws_bal['G2'] = 'Current Year'
ws_bal['H2'] = 'Prior Year'
for c in ['C2', 'D2', 'G2', 'H2']:
    ws_bal[c].fill = header_fill
    ws_bal[c].font = header_font

assets = [
    ('Current Assets', False),
    ('Cash and marketable securities', True),
    ('Receivables', True),
    ('Inventories', True),
    ('Other current assets', True),
    ('   Total current assets', False),
    ('', False),
    ('Fixed Assets', False),
    ('Property, plant, and equipment', True),
    ('Less accumulated depreciation', True),
    ('   Net tangible fixed assets', False),
    ('', False),
    ('Other Assets', False),
    ('Intangible assets (goodwill)', True),
    ('Other assets', True),
    ('', False),
    ('Total Assets', False),
]

for i, (label, is_input) in enumerate(assets, start=3):
    ws_bal.cell(row=i, column=2, value=label)
    if label and 'Total' in label:
        ws_bal.cell(row=i, column=2).font = bold
        ws_bal.cell(row=i, column=3).fill = green_fill
        ws_bal.cell(row=i, column=4).fill = green_fill
    elif is_input:
        ws_bal.cell(row=i, column=3).fill = yellow_fill
        ws_bal.cell(row=i, column=4).fill = yellow_fill

liabilities = [
    ('Current Liabilities', False),
    ('Debt due for repayment', True),
    ('Accounts payable', True),
    ('Other current liabilities', True),
    ('   Total current liabilities', False),
    ('', False),
    ('Long-term debt', True),
    ('Other long-term liabilities', True),
    ('', False),
    ('Total Liabilities', False),
    ('', False),
    ("SHAREHOLDERS' EQUITY", False),
    ('Common stock and paid-in capital', True),
    ('Retained earnings', True),
    ("   Total shareholders' equity", False),
    ('', False),
    ('Total Liabilities + Equity', False),
]

for i, (label, is_input) in enumerate(liabilities, start=3):
    ws_bal.cell(row=i, column=6, value=label)
    if label and 'Total' in label:
        ws_bal.cell(row=i, column=6).font = bold
        ws_bal.cell(row=i, column=7).fill = green_fill
        ws_bal.cell(row=i, column=8).fill = green_fill
    elif is_input:
        ws_bal.cell(row=i, column=7).fill = yellow_fill
        ws_bal.cell(row=i, column=8).fill = yellow_fill

# ========== INCOME STATEMENT TAB ==========
ws_inc = wb.create_sheet('Income Statement')
ws_inc.column_dimensions['A'].width = 50
ws_inc.column_dimensions['B'].width = 18
ws_inc.column_dimensions['C'].width = 15

ws_inc['A1'] = 'Income Statement'
ws_inc['A1'].font = section_font
ws_inc['B1'] = 'Amount'
ws_inc['C1'] = '% of Sales'
ws_inc['B1'].fill = header_fill
ws_inc['B1'].font = header_font
ws_inc['C1'].fill = header_fill
ws_inc['C1'].font = header_font

inc_items = [
    ('Net sales', 'yellow'),
    ('less Cost of goods sold', 'yellow'),
    ('less Selling, general & administrative expenses', 'yellow'),
    ('less Depreciation', 'yellow'),
    ('   Earnings before interest and taxes (EBIT)', 'green'),
    ('', ''),
    ('plus Other income', 'yellow'),
    ('less Interest expense', 'yellow'),
    ('  Taxable income', 'green'),
    ('', ''),
    ('less Taxes', 'yellow'),
    ('   Net income', 'green'),
    ('', ''),
    ('Allocation of net income', ''),
    ('   Dividends', 'yellow'),
    ('   Addition to retained earnings', 'green'),
]

for i, (label, color) in enumerate(inc_items, start=2):
    ws_inc.cell(row=i, column=1, value=label)
    if color == 'yellow':
        ws_inc.cell(row=i, column=2).fill = yellow_fill
    elif color == 'green':
        ws_inc.cell(row=i, column=2).fill = green_fill
        ws_inc.cell(row=i, column=1).font = bold

# ========== CASH FLOW STATEMENT TAB ==========
ws_cf = wb.create_sheet('Cash Flow Statement')
ws_cf.column_dimensions['A'].width = 55
ws_cf.column_dimensions['B'].width = 18
ws_cf.column_dimensions['C'].width = 20

ws_cf['A1'] = 'Cash Flow Statement'
ws_cf['A1'].font = section_font
ws_cf['B1'] = 'Amount'
ws_cf['B1'].fill = header_fill
ws_cf['B1'].font = header_font
ws_cf['C1'] = 'Source'
ws_cf['C1'].fill = header_fill
ws_cf['C1'].font = header_font

cf_items = [
    ('Operations:', '', ''),
    ('Net income', 'yellow', '[Income Statement]'),
    ('plus Depreciation', 'yellow', '[Income Statement]'),
    ('   plus Changes in working capital items', '', ''),
    ('      Decrease (increase) in accounts receivable', 'yellow', '[Balance Sheet]'),
    ('      Decrease (increase) in inventories', 'yellow', '[Balance Sheet]'),
    ('      Decrease (increase) in other current assets', 'yellow', '[Balance Sheet]'),
    ('      Increase (decrease) in accounts payable', 'yellow', '[Balance Sheet]'),
    ('      Increase (decrease) in other current liabilities', 'yellow', '[Balance Sheet]'),
    ('         Total change in working capital', 'green', ''),
    ('            Cash provided by operations', 'green', ''),
    ('', '', ''),
    ('Investments:', '', ''),
    ('Capital expenditures', 'yellow', ''),
    ('plus Sales (acquisitions) of long-term assets', 'yellow', ''),
    ('plus Other investing activities', 'yellow', ''),
    ('   Cash provided by (used for) investments', 'green', ''),
    ('', '', ''),
    ('Financing activities:', '', ''),
    ('Increase (decrease) in short-term debt', 'yellow', ''),
    ('plus Increase (decrease) in long-term debt', 'yellow', ''),
    ('plus Dividends', 'yellow', ''),
    ('plus Issues (repurchases) of stock', 'yellow', ''),
    ('plus Other', 'yellow', ''),
    ('   Cash provided by (used for) financing', 'green', ''),
    ('', '', ''),
    ('Net increase (decrease) in cash', 'green', ''),
]

for i, (label, color, source) in enumerate(cf_items, start=2):
    ws_cf.cell(row=i, column=1, value=label)
    ws_cf.cell(row=i, column=3, value=source)
    if color == 'yellow':
        ws_cf.cell(row=i, column=2).fill = yellow_fill
    elif color == 'green':
        ws_cf.cell(row=i, column=2).fill = green_fill
        ws_cf.cell(row=i, column=1).font = bold

# ========== RATIOS TAB ==========
ws_rat = wb.create_sheet('Ratios')
ws_rat.column_dimensions['A'].width = 48
ws_rat.column_dimensions['B'].width = 18
ws_rat.column_dimensions['C'].width = 65
ws_rat.column_dimensions['D'].width = 40

ws_rat['A1'] = 'Metric'
ws_rat['B1'] = 'Input'
ws_rat['C1'] = 'Named Range Formula'
ws_rat['D1'] = 'Named Range'
for c in ['A1', 'B1', 'C1', 'D1']:
    ws_rat[c].fill = header_fill
    ws_rat[c].font = header_font

inputs_section = [
    ('Share price', 'blue', '', 'share_price'),
    ('Shares outstanding (millions)', 'blue', '', 'shares_outstanding'),
    ('Cost of capital', 'blue', '', 'cost_capital'),
    ('Tax rate', 'blue', '', 'tax_rate'),
    ('Market capitalization', 'green', 'share_price * shares_outstanding', 'market_capitalization'),
    ('', '', '', ''),
    ('Start of Year (Prior Year Balance Sheet)', 'section', '', ''),
    ('Equity (start of year)', 'green', 'BAL_equity_shareholders_[prior_year]', 'startYear_equity'),
    ('Inventories (start of year)', 'green', 'BAL_inventories_[prior_year]', 'startYear_inventory'),
    ('Receivables (start of year)', 'green', 'BAL_receivables_[prior_year]', 'startYear_receivables'),
    ('Total assets (start of year)', 'green', 'BAL_assets_total_[prior_year]', 'startYear_total_assets'),
    ('Total capitalization (start of year)', 'green', 'BAL_debt_long_term_[prior] + BAL_equity_[prior]', 'startYear_total_capitalization'),
    ('', '', '', ''),
    ('Current Year (Derived)', 'section', '', ''),
    ('After-tax operating income', 'green', 'INC_net + (1 - tax_rate) * INC_interest_expense', 'currentYear_after_tax_operating_income'),
    ('Average daily sales', 'green', 'INC_sales / 365', 'currentYear_daily_sales_average'),
    ('Book value of equity', 'green', 'BAL_equity_shareholders_[current_year]', 'currentYear_equity'),
    ('Cash + Marketable securities', 'green', 'BAL_cash_marketable_securities_[current_year]', 'currentYear_cash_marketable_securities'),
    ('Current assets', 'green', 'BAL_assets_current_[current_year]', 'currentYear_assets_current'),
    ('Current liabilities', 'green', 'BAL_liabilities_current_[current_year]', 'currentYear_liabilities_current'),
    ('Daily COGS', 'green', 'INC_cost_goods_sold / 365', 'currentYear_cost_goods_sold_daily'),
    ('Long-term debt', 'green', 'BAL_debt_long_term_[current_year]', 'currentYear_debt_long_term'),
    ('Net working capital', 'green', 'currentYear_assets_current - currentYear_liabilities_current', 'currentYear_working_capital_net'),
    ('Total assets', 'green', 'BAL_assets_total_[current_year]', 'currentYear_assets_total'),
    ('Total capitalization', 'green', 'currentYear_debt_long_term + currentYear_equity', 'currentYear_total_capitalization'),
    ('Total liabilities', 'green', 'BAL_liabilities_total_[current_year]', 'currentYear_liabilities_total'),
    ('', '', '', ''),
    ('Mixed Year (Averages)', 'section', '', ''),
    ('Average equity', 'green', 'AVERAGE(startYear_equity, currentYear_equity)', 'avg_equity'),
    ('Average total assets', 'green', 'AVERAGE(startYear_total_assets, currentYear_assets_total)', 'avg_total_assets'),
    ('Average total capitalization', 'green', 'AVERAGE(startYear_total_capitalization, currentYear_total_capitalization)', 'avg_total_capitalization'),
    ('Long-term debt + Equity', 'green', 'currentYear_debt_long_term + currentYear_equity', 'debt_long_term_equity'),
]

row = 2
for label, color, formula, named_range in inputs_section:
    ws_rat.cell(row=row, column=1, value=label)
    ws_rat.cell(row=row, column=3, value=formula)
    ws_rat.cell(row=row, column=4, value=named_range)
    if color == 'blue':
        ws_rat.cell(row=row, column=2).fill = blue_fill
    elif color == 'green':
        ws_rat.cell(row=row, column=2).fill = green_fill
    elif color == 'section':
        ws_rat.cell(row=row, column=1).font = bold
    row += 1

row += 1
ws_rat.cell(row=row, column=1, value='Ratio')
ws_rat.cell(row=row, column=2, value='Output')
ws_rat.cell(row=row, column=3, value='Named Range Formula')
ws_rat.cell(row=row, column=4, value='Named Range')
for c in range(1, 5):
    ws_rat.cell(row=row, column=c).fill = header_fill
    ws_rat.cell(row=row, column=c).font = header_font
row += 1

ratios = [
    ('Performance', 'section', '', ''),
    ('Market value added (MVA)', 'gray', 'market_capitalization - currentYear_equity', ''),
    ('Market-to-book ratio', 'gray', 'market_capitalization / currentYear_equity', ''),
    ('Economic value added (EVA)', 'gray', 'currentYear_after_tax_operating_income - (cost_capital * startYear_total_capitalization)', ''),
    ('', '', '', ''),
    ('Profitability', 'section', '', ''),
    ('Return on assets (ROA)', 'gray', 'currentYear_after_tax_operating_income / startYear_total_assets', ''),
    ('Return on capital (ROC)', 'gray', 'currentYear_after_tax_operating_income / startYear_total_capitalization', ''),
    ('Return on equity (ROE)', 'gray', 'INC_net / startYear_equity', ''),
    ('Return on assets (ROA) [AVG]', 'gray', 'currentYear_after_tax_operating_income / avg_total_assets', ''),
    ('Return on capital (ROC) [AVG]', 'gray', 'currentYear_after_tax_operating_income / avg_total_capitalization', ''),
    ('Return on equity (ROE) [AVG]', 'gray', 'INC_net / avg_equity', ''),
    ('', '', '', ''),
    ('Efficiency', 'section', '', ''),
    ('Asset turnover', 'gray', 'INC_sales / startYear_total_assets', 'RATIO_asset_turnover'),
    ('Receivables turnover', 'gray', 'INC_sales / startYear_receivables', ''),
    ('Average collection period (days)', 'gray', 'startYear_receivables / currentYear_daily_sales_average', ''),
    ('Inventory turnover', 'gray', 'INC_cost_goods_sold / startYear_inventory', ''),
    ('Days in inventory', 'gray', 'startYear_inventory / currentYear_cost_goods_sold_daily', ''),
    ('Profit margin', 'gray', 'INC_net / INC_sales', ''),
    ('Operating profit margin', 'gray', 'currentYear_after_tax_operating_income / INC_sales', 'RATIO_operating_profit_margin'),
    ('', '', '', ''),
    ('Leverage', 'section', '', ''),
    ('Long-term debt ratio', 'gray', 'currentYear_debt_long_term / (currentYear_debt_long_term + currentYear_equity)', ''),
    ('Long-term debt-equity ratio', 'gray', 'currentYear_debt_long_term / currentYear_equity', ''),
    ('Total debt ratio', 'gray', 'currentYear_liabilities_total / currentYear_assets_total', ''),
    ('Times interest earned', 'gray', 'INC_ebit / INC_interest_expense', ''),
    ('Cash coverage ratio', 'gray', '(INC_ebit + INC_depreciation) / INC_interest_expense', ''),
    ('Debt burden', 'gray', 'INC_net / currentYear_after_tax_operating_income', 'RATIO_debt_burden'),
    ('Leverage ratio', 'gray', 'currentYear_assets_total / currentYear_equity', 'RATIO_leverage'),
    ('', '', '', ''),
    ('Liquidity', 'section', '', ''),
    ('Net working capital to assets', 'gray', 'currentYear_working_capital_net / currentYear_assets_total', ''),
    ('Current ratio', 'gray', 'currentYear_assets_current / currentYear_liabilities_current', ''),
    ('Quick ratio', 'gray', '(currentYear_cash + BAL_receivables) / currentYear_liabilities_current', ''),
    ('Cash ratio', 'gray', 'currentYear_cash_marketable_securities / currentYear_liabilities_current', ''),
    ('', '', '', ''),
    ('Du Pont System', 'section', '', ''),
    ('Return on assets (Du Pont)', 'gray', 'RATIO_asset_turnover * RATIO_operating_profit_margin', ''),
    ('Return on equity (Du Pont)', 'gray', 'RATIO_leverage * RATIO_asset_turnover * RATIO_operating_profit_margin * RATIO_debt_burden', ''),
]

for label, color, formula, named_range in ratios:
    ws_rat.cell(row=row, column=1, value=label)
    ws_rat.cell(row=row, column=3, value=formula)
    ws_rat.cell(row=row, column=4, value=named_range)
    if color == 'gray':
        ws_rat.cell(row=row, column=2).fill = gray_fill
    elif color == 'section':
        ws_rat.cell(row=row, column=1).font = section_font
    row += 1

# ========== NOTES TAB ==========
ws_notes = wb.create_sheet('Notes')
ws_notes.column_dimensions['A'].width = 60
ws_notes.column_dimensions['B'].width = 15
ws_notes['A1'] = 'Notes & Assumptions'
ws_notes['A1'].font = section_font
ws_notes['A3'] = 'Company: [Your Company Name]'
ws_notes['A4'] = 'Fiscal Year: [YYYY]'
ws_notes['A5'] = 'Data Source: [SEC EDGAR 10-K / Yahoo Finance / etc.]'
ws_notes['A6'] = 'Units: All figures in millions unless otherwise noted.'
ws_notes['A7'] = 'Tax Rate: [Statutory 21% / Effective rate from financials]'
ws_notes['A8'] = 'Cost of Capital: [X.X% - source/method]'
ws_notes['A10'] = 'Color Coding:'
ws_notes['A11'] = 'Yellow = Editable input cells'
ws_notes['B11'] = ''
ws_notes['B11'].fill = yellow_fill
ws_notes['A12'] = 'Blue = Assumptions'
ws_notes['B12'] = ''
ws_notes['B12'].fill = blue_fill
ws_notes['A13'] = 'Green = Formula cells'
ws_notes['B13'] = ''
ws_notes['B13'].fill = green_fill
ws_notes['A14'] = 'Gray = Output / summary cells'
ws_notes['B14'] = ''
ws_notes['B14'].fill = gray_fill
ws_notes['A16'] = 'AI & LLM Usage:'
ws_notes['A17'] = '[Document any AI tools used and how they were used]'

import os
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Stage3_Skeleton.xlsx')
wb.save(output_path)
print(f'Saved to {output_path}')
print('Done!')
